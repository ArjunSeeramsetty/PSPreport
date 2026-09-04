"""Synthetic WBES documents. These are not captured Grid-India payloads."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook

from psp_pipeline.wbes.blocks import DEFAULT_BLOCK_COUNT


def block_values(base: float, *, count: int = DEFAULT_BLOCK_COUNT) -> list[float]:
    """Return a deterministic 96-block MW vector."""

    return [round(base + (index / 100.0), 4) for index in range(count)]


def synthetic_document(
    *,
    schedule_date: str = "2026-09-01",
    revision_label: str = "R0",
    revision_no: int | None = None,
    source_region: str = "WR",
    block_count: int = DEFAULT_BLOCK_COUNT,
    injection_base: float = 200.0,
    requisition_split: tuple[float, float] = (120.0, 80.0),
) -> dict[str, object]:
    """Build a tiny canonical entitlement/requisition/net-schedule document."""

    isgs = "SYNTH-ISGS-1"
    state_a = "SYNTH-STATE-A"
    state_b = "SYNTH-STATE-B"
    tie = "SYNTH-TIE-WR-NR"
    req_a, req_b = requisition_split
    payload: dict[str, object] = {
        "schedule_date": schedule_date,
        "revision_label": revision_label,
        "source_region": source_region,
        "source_id": "wbes_national",
        "block_count": block_count,
        "block_minutes": 15 if block_count == 96 else 5,
        "matrices": [
            {
                "kind": "entitlement",
                "rows": [
                    {
                        "entity_id": isgs,
                        "entity_name": "Synthetic ISGS 1",
                        "archetype": "isgs",
                        "counterparty_id": state_a,
                        "counterparty_name": "Synthetic State A",
                        "counterparty_archetype": "beneficiary",
                        "values_mw": block_values(150.0, count=block_count),
                    },
                    {
                        "entity_id": isgs,
                        "entity_name": "Synthetic ISGS 1",
                        "archetype": "isgs",
                        "counterparty_id": state_b,
                        "counterparty_name": "Synthetic State B",
                        "counterparty_archetype": "beneficiary",
                        "values_mw": block_values(50.0, count=block_count),
                    },
                ],
            },
            {
                "kind": "requisition",
                "rows": [
                    {
                        "entity_id": isgs,
                        "entity_name": "Synthetic ISGS 1",
                        "archetype": "isgs",
                        "counterparty_id": state_a,
                        "counterparty_name": "Synthetic State A",
                        "counterparty_archetype": "beneficiary",
                        "values_mw": block_values(req_a, count=block_count),
                    },
                    {
                        "entity_id": isgs,
                        "entity_name": "Synthetic ISGS 1",
                        "archetype": "isgs",
                        "counterparty_id": state_b,
                        "counterparty_name": "Synthetic State B",
                        "counterparty_archetype": "beneficiary",
                        "values_mw": block_values(req_b, count=block_count),
                    },
                ],
            },
            {
                "kind": "net_schedule",
                "component": "injection",
                "rows": [
                    {
                        "entity_id": isgs,
                        "entity_name": "Synthetic ISGS 1",
                        "archetype": "isgs",
                        "values_mw": block_values(injection_base, count=block_count),
                    },
                    {
                        "entity_id": tie,
                        "entity_name": "Synthetic WR-NR tie",
                        "archetype": "regional_tie",
                        "values_mw": block_values(25.0, count=block_count),
                    },
                ],
            },
            {
                "kind": "net_schedule",
                "component": "drawal",
                "rows": [
                    {
                        "entity_id": state_a,
                        "entity_name": "Synthetic State A",
                        "archetype": "beneficiary",
                        "values_mw": block_values(req_a, count=block_count),
                    }
                ],
            },
            {
                "kind": "net_schedule",
                "component": "bilateral",
                "rows": [
                    {
                        "entity_id": state_a,
                        "entity_name": "Synthetic State A",
                        "archetype": "beneficiary",
                        "values_mw": block_values(10.0, count=block_count),
                    }
                ],
            },
            {
                "kind": "net_schedule",
                "component": "collective",
                "rows": [
                    {
                        "entity_id": state_a,
                        "entity_name": "Synthetic State A",
                        "archetype": "beneficiary",
                        "values_mw": block_values(5.0, count=block_count),
                    }
                ],
            },
        ],
    }
    if revision_no is not None:
        payload["revision_no"] = revision_no
    return payload


def write_json(path: Path, payload: dict[str, object]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload), encoding="utf-8")
    return path


def write_xlsx(path: Path, payload: dict[str, object]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    meta = workbook.active
    meta.title = "meta"
    meta.append(["schedule_date", payload["schedule_date"]])
    meta.append(["revision_label", payload["revision_label"]])
    meta.append(["source_region", payload["source_region"]])
    meta.append(["source_id", payload["source_id"]])
    meta.append(["block_count", payload["block_count"]])
    meta.append(["block_minutes", payload["block_minutes"]])
    block_count = int(payload["block_count"])
    header = [
        "entity_id",
        "entity_name",
        "archetype",
        "counterparty_id",
        "counterparty_name",
        "counterparty_archetype",
        *[str(block_no) for block_no in range(1, block_count + 1)],
    ]
    for matrix in payload["matrices"]:
        sheet_name = str(matrix["kind"])
        if matrix.get("component"):
            sheet_name = f"{sheet_name}_{matrix['component']}"
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)
            sheet.append(header)
        for row in matrix["rows"]:
            sheet.append(
                [
                    row["entity_id"],
                    row["entity_name"],
                    row["archetype"],
                    row.get("counterparty_id"),
                    row.get("counterparty_name"),
                    row.get("counterparty_archetype"),
                    *row["values_mw"],
                ]
            )
    workbook.save(path)
    return path
