"""Extract RPC settlement tables from local PDF or Excel artifacts."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from psp_pipeline.parsing.rldc.pdf_tables import extract_page_tables


@dataclass(frozen=True)
class ExtractedTable:
    """One physical table retained with stable page/sheet and table indexes."""

    page_no: int
    table_no: int
    sheet_name: str | None
    rows: tuple[tuple[str, ...], ...]


def extract_rpc_tables(path: Path | str) -> tuple[ExtractedTable, ...]:
    """Return local PDF or Excel tables without contacting the source portal.

    Excel sheets are numbered from one so lineage can treat a sheet as a page.
    Empty sheets and decorative one-cell fragments are dropped.
    """

    local = Path(path)
    suffix = local.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xls"}:
        return _extract_excel_tables(local)
    return _extract_pdf_tables(local)


def table_as_column_maps(
    table: ExtractedTable,
) -> list[dict[int, tuple[int, str]]]:
    """Project extracted rows into the 1-based raw-cell maps used by promoters."""

    projected: list[dict[int, tuple[int, str]]] = []
    for row_no, row in enumerate(table.rows, start=1):
        mapped: dict[int, tuple[int, str]] = {}
        for col_no, text in enumerate(row, start=1):
            if str(text or "").strip():
                mapped[col_no] = (0, str(text))
        if mapped:
            projected.append(mapped)
    return projected


def _extract_pdf_tables(path: Path) -> tuple[ExtractedTable, ...]:
    """Extract native or text-strategy tables from a local RPC PDF."""

    import pdfplumber

    tables: list[ExtractedTable] = []
    with pdfplumber.open(str(path)) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):
            for table_no, raw in enumerate(extract_page_tables(page), start=1):
                rows = _normalize_matrix(raw)
                if rows:
                    tables.append(
                        ExtractedTable(page_no, table_no, None, rows)
                    )
    return tuple(tables)


def _extract_excel_tables(path: Path) -> tuple[ExtractedTable, ...]:
    """Extract used ranges from each worksheet as one table per sheet."""

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=False)
    tables: list[ExtractedTable] = []
    try:
        for page_no, sheet in enumerate(workbook.worksheets, start=1):
            rows = _normalize_matrix(
                tuple(
                    tuple("" if cell is None else str(cell) for cell in row)
                    for row in sheet.iter_rows(values_only=True)
                )
            )
            if rows:
                tables.append(
                    ExtractedTable(page_no, 1, sheet.title, rows)
                )
    finally:
        workbook.close()
    return tuple(tables)


def _normalize_matrix(raw: object) -> tuple[tuple[str, ...], ...]:
    """Strip trailing empty rows/columns while preserving column alignment."""

    rows = [
        tuple("" if cell is None else str(cell).strip() for cell in row)
        for row in (raw or ())
    ]
    while rows and not any(cell.strip() for cell in rows[-1]):
        rows.pop()
    if not rows:
        return ()
    width = max(len(row) for row in rows)
    padded = [row + ("",) * (width - len(row)) for row in rows]
    while width and all(not row[width - 1].strip() for row in padded):
        width -= 1
        padded = [row[:width] for row in padded]
    populated = sum(1 for row in padded for cell in row if cell.strip())
    if len(padded) < 2 or width < 3 or populated < 6:
        return ()
    return tuple(padded)
