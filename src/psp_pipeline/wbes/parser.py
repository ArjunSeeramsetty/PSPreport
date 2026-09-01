"""Parse dropped or captured WBES schedule matrices into revision documents."""

from __future__ import annotations

from collections.abc import Iterable
from datetime import date, datetime, timezone
import hashlib
import json
from pathlib import Path
from typing import Any

from psp_pipeline.wbes.blocks import require_standard_blocks, validate_value_vector
from psp_pipeline.wbes.models import (
    EntityArchetype,
    MatrixKind,
    ScheduleComponent,
    WbesMatrix,
    WbesMatrixRow,
    WbesRevisionDocument,
    parse_revision_label,
)

PAIR_KINDS = {MatrixKind.ENTITLEMENT, MatrixKind.REQUISITION}


class WbesParseError(ValueError):
    """Raised when a drop file is not a usable WBES schedule matrix."""


def parse_wbes_path(
    path: Path,
    *,
    source_id: str,
    source_region: str,
    block_count: int,
    block_minutes: int,
    allow_five_minute: bool,
    source_url: str = "",
) -> WbesRevisionDocument:
    """Parse a canonical JSON or wide XLSX drop file."""

    require_standard_blocks(
        block_count=block_count,
        minutes=block_minutes,
        allow_five_minute=allow_five_minute,
    )
    suffix = path.suffix.lower()
    if suffix == ".json":
        payload = json.loads(path.read_text(encoding="utf-8"))
        document = parse_wbes_payload(
            payload,
            source_id=source_id,
            source_region=source_region,
            block_count=block_count,
            block_minutes=block_minutes,
            source_url=source_url,
            raw_path=str(path),
        )
    elif suffix in {".xlsx", ".xlsm"}:
        payload = _xlsx_to_payload(path)
        document = parse_wbes_payload(
            payload,
            source_id=source_id,
            source_region=source_region,
            block_count=block_count,
            block_minutes=block_minutes,
            source_url=source_url,
            raw_path=str(path),
        )
    else:
        raise WbesParseError(f"Unsupported WBES drop format: {path.suffix}")
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    return _with_hash(document, digest)


def parse_wbes_payload(
    payload: dict[str, Any],
    *,
    source_id: str,
    source_region: str,
    block_count: int,
    block_minutes: int,
    source_url: str = "",
    raw_path: str = "",
    content_hash: str = "",
) -> WbesRevisionDocument:
    """Parse an in-memory canonical schedule document."""

    schedule_date = date.fromisoformat(str(payload["schedule_date"]))
    revision_label, parsed_revision = parse_revision_label(str(payload.get("revision_label", "R0")))
    raw_revision = payload.get("revision_no")
    revision_no = int(raw_revision) if raw_revision is not None and str(raw_revision) != "" else parsed_revision
    declared_blocks = int(payload.get("block_count", block_count))
    declared_minutes = int(payload.get("block_minutes", block_minutes))
    if declared_blocks != block_count or declared_minutes != block_minutes:
        raise WbesParseError(
            f"document grain {declared_blocks}x{declared_minutes}m does not match "
            f"pipeline grain {block_count}x{block_minutes}m"
        )
    region = str(payload.get("source_region", source_region))
    matrices = tuple(
        _parse_matrix(item, block_count=block_count) for item in payload.get("matrices", [])
    )
    if not matrices:
        raise WbesParseError("WBES document contains no matrices")
    published_at = _parse_optional_datetime(payload.get("published_at"))
    document = WbesRevisionDocument(
        schedule_date=schedule_date,
        revision_label=revision_label,
        revision_no=revision_no,
        source_region=region,
        source_id=str(payload.get("source_id", source_id)),
        block_count=block_count,
        block_minutes=block_minutes,
        matrices=matrices,
        published_at=published_at,
        content_hash=content_hash,
        source_url=source_url,
        raw_path=raw_path,
    )
    return document


def iter_drop_files(drop_dir: Path) -> tuple[Path, ...]:
    """Return JSON/XLSX files from the isolated drop directory."""

    if not drop_dir.exists():
        return ()
    files = [
        path
        for path in drop_dir.rglob("*")
        if path.is_file() and path.suffix.lower() in {".json", ".xlsx", ".xlsm"}
    ]
    return tuple(sorted(files))


def _parse_matrix(item: dict[str, Any], *, block_count: int) -> WbesMatrix:
    kind = MatrixKind(str(item["kind"]).lower())
    component = item.get("component")
    parsed_component = ScheduleComponent(str(component).lower()) if component else None
    rows = tuple(_parse_row(row, kind=kind, block_count=block_count) for row in item.get("rows", []))
    if not rows:
        raise WbesParseError(f"{kind.value} matrix has no rows")
    return WbesMatrix(kind=kind, component=parsed_component, rows=rows)


def _parse_row(row: dict[str, Any], *, kind: MatrixKind, block_count: int) -> WbesMatrixRow:
    try:
        values = _coerce_values(row, block_count=block_count)
        validate_value_vector(values, block_count=block_count)
    except ValueError as exc:
        raise WbesParseError(str(exc)) from exc
    archetype = EntityArchetype(str(row.get("archetype", _default_archetype(kind))).lower())
    counterparty_id = row.get("counterparty_id")
    if kind in PAIR_KINDS and not counterparty_id:
        raise WbesParseError(f"{kind.value} rows require counterparty_id")
    counterpart_archetype = row.get("counterparty_archetype")
    return WbesMatrixRow(
        entity_id=str(row["entity_id"]),
        entity_name=str(row.get("entity_name", row["entity_id"])),
        archetype=archetype,
        values_mw=tuple(values),
        counterparty_id=str(counterparty_id) if counterparty_id else None,
        counterparty_name=(
            str(row.get("counterparty_name", counterparty_id)) if counterparty_id else None
        ),
        counterparty_archetype=(
            EntityArchetype(str(counterpart_archetype).lower())
            if counterpart_archetype
            else (EntityArchetype.BENEFICIARY if counterparty_id else None)
        ),
    )


def _default_archetype(kind: MatrixKind) -> str:
    if kind is MatrixKind.NET_SCHEDULE:
        return EntityArchetype.ISGS.value
    return EntityArchetype.ISGS.value


def _coerce_values(row: dict[str, Any], *, block_count: int) -> list[float]:
    if "values_mw" in row:
        return [float(value) for value in row["values_mw"]]
    blocks = row.get("blocks")
    if isinstance(blocks, dict):
        values: list[float] = []
        for block_no in range(1, block_count + 1):
            if str(block_no) not in blocks and block_no not in blocks:
                raise WbesParseError(f"missing block {block_no}")
            values.append(float(blocks.get(str(block_no), blocks.get(block_no))))
        return values
    raise WbesParseError("row is missing values_mw or blocks")


def _parse_optional_datetime(value: object) -> datetime | None:
    if not value:
        return None
    parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=timezone.utc)
    return parsed


def _with_hash(document: WbesRevisionDocument, digest: str) -> WbesRevisionDocument:
    return WbesRevisionDocument(
        schedule_date=document.schedule_date,
        revision_label=document.revision_label,
        revision_no=document.revision_no,
        source_region=document.source_region,
        source_id=document.source_id,
        block_count=document.block_count,
        block_minutes=document.block_minutes,
        matrices=document.matrices,
        published_at=document.published_at,
        content_hash=digest,
        source_url=document.source_url,
        raw_path=document.raw_path,
    )


def _xlsx_to_payload(path: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - optional extra
        raise WbesParseError("openpyxl is required to parse WBES XLSX drop files") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        names = {name.lower(): name for name in workbook.sheetnames}
        if "meta" not in names:
            raise WbesParseError("XLSX drop files need a meta sheet with schedule_date")
        meta = _read_meta_sheet(workbook[names["meta"]])
        payload = {
            "schedule_date": _as_iso_date(meta.get("schedule_date")),
            "revision_label": meta.get("revision_label", "R0"),
            "revision_no": meta.get("revision_no"),
            "source_region": meta.get("source_region", "ALL"),
            "source_id": meta.get("source_id", "wbes_national"),
            "block_count": int(meta.get("block_count") or 96),
            "block_minutes": int(meta.get("block_minutes") or 15),
            "matrices": [],
        }
        if payload["schedule_date"] is None:
            raise WbesParseError("XLSX drop files need a meta sheet with schedule_date")
        matrices: list[dict[str, Any]] = []
        for original in workbook.sheetnames:
            if original.lower() == "meta":
                continue
            kind, component = _sheet_kind(original)
            rows = _xlsx_rows(workbook[original], block_count=int(payload["block_count"]))
            matrices.append(
                {
                    "kind": kind.value,
                    "component": component.value if component else None,
                    "rows": rows,
                }
            )
        payload["matrices"] = matrices
        return payload
    finally:
        workbook.close()


def _read_meta_sheet(sheet: Any) -> dict[str, Any]:
    values: dict[str, Any] = {}
    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        if row[0]:
            values[str(row[0]).strip()] = row[1]
    return values


def _as_iso_date(value: object) -> str | None:
    if value is None:
        return None
    if hasattr(value, "date") and callable(getattr(value, "date")):
        return value.date().isoformat()
    if hasattr(value, "isoformat") and not isinstance(value, str):
        return str(value.isoformat())[:10]
    text = str(value).strip()
    return text[:10] if text else None


def _sheet_kind(name: str) -> tuple[MatrixKind, ScheduleComponent | None]:
    lowered = name.strip().lower().replace(" ", "_")
    if lowered.startswith("net_schedule_"):
        return MatrixKind.NET_SCHEDULE, ScheduleComponent(lowered.split("net_schedule_", 1)[1])
    if lowered == "net_schedule":
        return MatrixKind.NET_SCHEDULE, None
    return MatrixKind(lowered), None


def _xlsx_rows(sheet: Any, *, block_count: int) -> list[dict[str, Any]]:
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = [str(cell).strip() if cell is not None else "" for cell in next(rows_iter)]
    except StopIteration as exc:
        raise WbesParseError("XLSX sheet is empty") from exc
    index = {name.lower(): offset for offset, name in enumerate(header) if name}
    rows: list[dict[str, Any]] = []
    for raw in rows_iter:
        if not raw or raw[index.get("entity_id", 0)] is None:
            continue
        values = []
        for block_no in range(1, block_count + 1):
            column = index.get(str(block_no)) or index.get(f"b{block_no}")
            if column is None:
                raise WbesParseError(f"XLSX sheet is missing block column {block_no}")
            values.append(float(raw[column] or 0.0))
        rows.append(
            {
                "entity_id": raw[index["entity_id"]],
                "entity_name": raw[index["entity_name"]] if "entity_name" in index else raw[index["entity_id"]],
                "archetype": raw[index["archetype"]] if "archetype" in index else "isgs",
                "counterparty_id": raw[index["counterparty_id"]] if "counterparty_id" in index else None,
                "counterparty_name": raw[index["counterparty_name"]] if "counterparty_name" in index else None,
                "counterparty_archetype": (
                    raw[index["counterparty_archetype"]] if "counterparty_archetype" in index else None
                ),
                "values_mw": values,
            }
        )
    return rows


def discover_schedule_dates(documents: Iterable[WbesRevisionDocument]) -> tuple[str, ...]:
    """Return sorted ISO dates present in parsed documents."""

    return tuple(sorted({item.schedule_date.isoformat() for item in documents}))
